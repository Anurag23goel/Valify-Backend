import os
import json
import tempfile
import requests
import base64
from datetime import datetime
import firebase_admin
from firebase_admin import credentials, firestore, storage, auth
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from dotenv import load_dotenv
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import fitz 
import shutil
import warnings

warnings.simplefilter("ignore", UserWarning)



# Load environment variables
if os.getenv("RENDER") is None:  # Render automatically sets this variable
    if not load_dotenv():
        print("Warning: .env file not found. Ensure environment variables are set!")

# Function to verify Firebase Authentication Token
def verify_token(id_token):
    try:
        decoded_token = auth.verify_id_token(id_token)
        return decoded_token
    except auth.ExpiredIdTokenError:
        raise Exception("Token has expired")
    except auth.RevokedIdTokenError:
        raise Exception("Token has been revoked")
    except Exception as e:
        raise Exception(f"Error verifying token: {str(e)}")

# Initialize Firebase
firebase_credentials = os.getenv("FIREBASE_CREDENTIALS")

if not firebase_credentials:
    raise ValueError("Missing FIREBASE_CREDENTIALS environment variable!")

try:
    cred_dict = json.loads(firebase_credentials)
    cred = credentials.Certificate(cred_dict)
    firebase_admin.initialize_app(cred, {"storageBucket": "valify-7e530.appspot.com"})
    db = firestore.client()
except json.JSONDecodeError as e:
    raise ValueError(f"Invalid FIREBASE_CREDENTIALS JSON: {e}")

CONVERT_API_KEY = os.getenv("CONVERT_API_KEY")
CONVERT_API_URL = "https://v2.convertapi.com/convert/xls/to/pdf"

if not CONVERT_API_KEY:
    raise ValueError("Missing CONVERT_API_KEY in environment variables!")

# Initialize Flask app
app = Flask(__name__)
CORS(app)

# Define paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "dynamic_excel.xlsx")  # Ensure this file exists
TEMPLATE_PATH_HIST = os.path.join(BASE_DIR, "hist_fin.xlsx")  # Ensure this file exists
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)


json_to_excel_mapping = {
    "Inputs": {
  "CashasatValuationDate": "E257",
  "CurrentAssets_0_0": "L232",
  "CurrentAssets_0_1": "M232",
  "CurrentAssets_0_2": "N232",
  "CurrentAssets_0_3": "O232",
  "CurrentAssets_0_4": "P232",
  "CurrentAssets_0_5": "Q232",
  "CurrentAssets_1_0": "L233",
  "CurrentAssets_1_1": "M233",
  "CurrentAssets_1_2": "N233",
  "CurrentAssets_1_3": "O233",
  "CurrentAssets_1_4": "P233",
  "CurrentAssets_1_5": "Q233",
  "CurrentAssets_2_0": "L234",
  "CurrentAssets_2_1": "M234",
  "CurrentAssets_2_2": "N234",
  "CurrentAssets_2_3": "O234",
  "CurrentAssets_2_4": "P234",
  "CurrentAssets_2_5": "Q234",
  "CurrentAssets_3_0": "L235",
  "CurrentAssets_3_1": "M235",
  "CurrentAssets_3_2": "N235",
  "CurrentAssets_3_3": "O235",
  "CurrentAssets_3_4": "P235",
  "CurrentAssets_3_5": "Q235",
  "Cyclicality": "E279",
  "PenetrationRisk": "E275",
  "VendorRisk": "E276",
  "avgAnnualRevenue": "E49",
  "clientName": "E19",
  "currentLiabilities_0_0": "L237",
  "currentLiabilities_0_1": "M237",
  "currentLiabilities_0_2": "N237",
  "currentLiabilities_0_3": "O237",
  "currentLiabilities_0_4": "P237",
  "currentLiabilities_0_5": "Q237",
  "currentLiabilities_1_0": "L238",
  "currentLiabilities_1_1": "M238",
  "currentLiabilities_1_2": "N238",
  "currentLiabilities_1_3": "O238",
  "currentLiabilities_1_4": "P238",
  "currentLiabilities_1_5": "Q238",
  "currentLiabilities_2_0": "L239",
  "currentLiabilities_2_1": "M239",
  "currentLiabilities_2_2": "N239",
  "currentLiabilities_2_3": "O239",
  "currentLiabilities_2_4": "P239",
  "currentLiabilities_2_5": "Q239",
  "currentLiabilities_3_0": "L240",
  "currentLiabilities_3_1": "M240",
  "currentLiabilities_3_2": "N240",
  "currentLiabilities_3_3": "O240",
  "currentLiabilities_3_4": "P240",
  "currentLiabilities_3_5": "Q240",
  "currentLiabilities_4_0": "L241",
  "currentLiabilities_4_1": "M241",
  "currentLiabilities_4_2": "N241",
  "currentLiabilities_4_3": "O241",
  "currentLiabilities_4_4": "P241",
  "currentLiabilities_4_5": "Q241",
  "developmentPhase": "E50",
  "existingStream1": "E77",
  "existingStream2": "E78",
  "existingStream3": "E79",
  "existingStream4": "E80",
  "existingStreamsGrossMargin_0_0": "L167",
  "existingStreamsGrossMargin_0_1": "M167",
  "existingStreamsGrossMargin_0_2": "N167",
  "existingStreamsGrossMargin_0_3": "O167",
  "existingStreamsGrossMargin_0_4": "P167",
  "existingStreamsGrossMargin_0_5": "Q167",
  "existingStreamsGrossMargin_1_0": "L168",
  "existingStreamsGrossMargin_1_1": "M168",
  "existingStreamsGrossMargin_1_2": "N168",
  "existingStreamsGrossMargin_1_3": "O168",
  "existingStreamsGrossMargin_1_4": "P168",
  "existingStreamsGrossMargin_1_5": "Q168",
  "existingStreamsGrossMargin_2_0": "L169",
  "existingStreamsGrossMargin_2_1": "M169",
  "existingStreamsGrossMargin_2_2": "N169",
  "existingStreamsGrossMargin_2_3": "O169",
  "existingStreamsGrossMargin_2_4": "P169",
  "existingStreamsGrossMargin_2_5": "Q169",
  "existingStreamsGrossMargin_3_0": "L170",
  "existingStreamsGrossMargin_3_1": "M170",
  "existingStreamsGrossMargin_3_2": "N170",
  "existingStreamsGrossMargin_3_3": "O170",
  "existingStreamsGrossMargin_3_4": "P170",
  "existingStreamsGrossMargin_3_5": "Q170",
  "industryPrimaryBusiness": "E52",
  "industrySecondaryBusiness": "E63",
  "informationCurrency": "E43",
  "otherOperatingRegionsSecondaryName1": "D71",
  "otherOperatingRegionsSecondaryName2": "D72",
  "otherOperatingRegionsSecondaryName3": "D73",
  "otherOperatingRegionsSecondaryValue1": "E71",
  "otherOperatingRegionsSecondaryValue2": "E72",
  "otherOperatingRegionsSecondaryValue3": "E73",
  "otherRegionsName1": "D59",
  "otherRegionsName2": "D60",
  "otherRegionsName3": "D61",
  "otherRegionsValue1": "E59",
  "otherRegionsValue2": "E60",
  "otherRegionsValue3": "E61",
  "pipelineStream1": "E82",
  "pipelineStream2": "E83",
  "pipelineStream3": "E84",
  "pipelineStream4": "E85",
  "pipelineStreamsGrossMargin_0_0": "L171",
  "pipelineStreamsGrossMargin_0_1": "M171",
  "pipelineStreamsGrossMargin_0_2": "N171",
  "pipelineStreamsGrossMargin_0_3": "O171",
  "pipelineStreamsGrossMargin_0_4": "P171",
  "pipelineStreamsGrossMargin_0_5": "Q171",
  "pipelineStreamsGrossMargin_1_0": "L172",
  "pipelineStreamsGrossMargin_1_1": "M172",
  "pipelineStreamsGrossMargin_1_2": "N172",
  "pipelineStreamsGrossMargin_1_3": "O172",
  "pipelineStreamsGrossMargin_1_4": "P172",
  "pipelineStreamsGrossMargin_1_5": "Q172",
  "pipelineStreamsGrossMargin_2_0": "L173",
  "pipelineStreamsGrossMargin_2_1": "M173",
  "pipelineStreamsGrossMargin_2_2": "N173",
  "pipelineStreamsGrossMargin_2_3": "O173",
  "pipelineStreamsGrossMargin_2_4": "P173",
  "pipelineStreamsGrossMargin_2_5": "Q173",
  "pipelineStreamsGrossMargin_3_0": "L174",
  "pipelineStreamsGrossMargin_3_1": "M174",
  "pipelineStreamsGrossMargin_3_2": "N174",
  "pipelineStreamsGrossMargin_3_3": "O174",
  "pipelineStreamsGrossMargin_3_4": "P174",
  "pipelineStreamsGrossMargin_3_5": "Q174",
  "potentialStream1": "E87",
  "potentialStream1Probability": "E93",
  "potentialStream2": "E88",
  "potentialStream2Probability": "E94",
  "potentialStream3": "E89",
  "potentialStream3Probability": "E95",
  "potentialStream4": "E90",
  "potentialStream4Probability": "E96",
  "presentationCurrency": "E44",
  "primaryBusiness": "E54",
  "premise": "E22",
  "draftNote": "E23",
  "projectTitle": "E26",
  "primaryBusinessDescription": "E55",
  "primaryRegions": "E56",
  "purpose": "E21",
  "secondaryRegions": "E67",
  "secondaryBusiness": "E65",
  "secondaryBusinessDescription": "E66",
  "shortName": "E25",
  "valuationDate": "E29",
  "nextFiscalYearEndDate": "E30",
  "subindustryPrimaryBusiness": "E53",
  "subindustrySecondaryBusiness": "E64",
  "subjectCompanyName": "E24",
  "units": "E45",
  "valuerName": "E20",
  "valuerType": "E18",
  "ytd": "E33",
  "ytgApproach": "E36"

    }
}
json_to_excel_mapping_currency = {
    "Hist.Fin": {
        "valuationDate": "D7",
        "informationCurrency": "D8",
        "units": "D9"
    }
}


# Route to remove formulas from an Excel file
@app.route('/remove-formulas', methods=['POST'])
def remove_formulas_route():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    excel_file = request.files['file']

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_input:
        excel_file.save(temp_input.name)
        
        # Define output file paths
        temp_output = temp_input.name.replace(".xlsx", "_no_formulas.xlsx")
        
        # Ensure output directory exists
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        output_filename = f"processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, output_filename)

        # Remove formulas and save the processed file
        remove_formulas_from_excel(temp_input.name, temp_output)

        # Move the file to the output directory
        os.rename(temp_output, output_path)

        return send_file(output_path, as_attachment=True)


# Route to generate an Excel file with Firestore data
@app.route('/generate-excel', methods=['GET'])
def generate_excel():
    """Flask route to generate an Excel file and return it as a response"""
    try:
        uid = request.args.get('uid')
        project_id = request.args.get('project_id')

        if not uid or not project_id:
            return jsonify({"error": "uid and project_id are required"}), 400

        # Call the function that generates Excel
        output_path = generate_excel_file(uid, project_id)

        return send_file(output_path, as_attachment=True)

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route('/generate-excel-hist', methods=['GET'])
def generate_excel_hist():
    """Flask route to generate an Excel file and return it as a response"""
    try:
        uid = request.args.get('uid')
        project_id = request.args.get('project_id')

        if not uid or not project_id:
            return jsonify({"error": "uid and project_id are required"}), 400

        # Call the function that generates Excel
        output_path = generate_excel_file_hist(uid, project_id)

        return send_file(output_path, as_attachment=True)

    except Exception as e:
        return jsonify({"error": str(e)}), 500



# Route to convert an Excel file to PDF
@app.route('/convert-to-pdf', methods=['GET'])
def convert_to_pdf_route():
    """GET route to generate an Excel file and convert it to a PDF"""
    uid = request.args.get("uid")
    project_id = request.args.get("project_id")

    if not uid or not project_id:
        return jsonify({"error": "Missing uid or project_id"}), 400

    try:
        # Generate the Excel file (without sending it as a response)
        excel_file_path = generate_excel_file(uid, project_id)
        if not os.path.exists(excel_file_path):
            return jsonify({"error": "Failed to generate Excel file"}), 500

        # Define PDF output path
        pdf_output_path = excel_file_path.replace(".xlsx", ".pdf")

        # Convert Excel to PDF
        conversion_success = convert_excel_to_pdf(excel_file_path, pdf_output_path)
        if not conversion_success:
            return jsonify({"error": "Failed to convert Excel to PDF"}), 500

        # Extract pages from 104 onward (optional step)
        extracted_pdf_output = pdf_output_path.replace(".pdf", "_report.pdf")
        extraction_success = extract_pages_from_pdf(pdf_output_path, extracted_pdf_output, start_page=104)

        if not extraction_success:
            return jsonify({"error": "Failed to extract report pages from PDF"}), 500

        # Store in output directory
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        final_pdf_path = os.path.join(OUTPUT_DIR, os.path.basename(extracted_pdf_output))
        shutil.move(extracted_pdf_output, final_pdf_path)

        # Return the final PDF file
        return send_file(final_pdf_path, as_attachment=True)

    except Exception as e:
        return jsonify({"error": str(e)}), 500



# Function to generate excel file
def generate_excel_file(uid: str, project_id: str) -> str:
    """Generates an Excel file with Firestore data and returns the file path"""
    try:
        if not uid or not project_id:
            raise ValueError("uid and project_id are required")

        # Fetch Firestore data
        doc_ref = db.collection("users").document(uid).collection("projects").document(project_id)
        doc = doc_ref.get()

        if not doc.exists:
            raise FileNotFoundError("Document not found")

        res = doc.to_dict()
        data = res.get("answers", {})

        # Load Excel template
        workbook = load_workbook(TEMPLATE_PATH, keep_vba=True, data_only=True)

        if "Inputs" not in workbook.sheetnames:
            raise Exception("Excel template is missing 'Inputs' sheet")

        worksheet = workbook["Inputs"]

        for field, cell_location in json_to_excel_mapping["Inputs"].items():
            value = data.get(field, None)
            if value is not None:
                worksheet[cell_location].value = value

        # Ensure output directory exists
        os.makedirs(OUTPUT_DIR, exist_ok=True)

        # Generate unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"final_invoice_{timestamp}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, output_filename)

        # Save and close the workbook
        workbook.save(output_path)
        workbook.close()

        # Ensure file exists before returning the path
        if not os.path.exists(output_path):
            raise Exception("Failed to generate Excel file")

        return output_path  # Return the file path

    except Exception as e:
        raise RuntimeError(f"Error generating Excel: {str(e)}")
    
# Function to generate excel file for historical template
def generate_excel_file_hist(uid: str, project_id: str) -> str:
    """Generates an Excel file with Firestore data and returns the file path"""
    try:
        if not uid or not project_id:
            raise ValueError("uid and project_id are required")

        # Fetch Firestore data
        doc_ref = db.collection("users").document(uid).collection("projects").document(project_id)
        doc = doc_ref.get()

        if not doc.exists:
            raise FileNotFoundError("Document not found")

        res = doc.to_dict()
        data = res.get("answers", {})

        # Load Excel template
        workbook = load_workbook(TEMPLATE_PATH_HIST, keep_vba=True, data_only=True)

        if "Hist.Fin" not in workbook.sheetnames:
            raise Exception("Excel template is missing 'Inputs' sheet")

        worksheet = workbook["Hist.Fin"]

        for field, cell_location in json_to_excel_mapping_currency["Hist.Fin"].items():
            value = data.get(field, None)
            if value is not None:
                worksheet[cell_location].value = value

        # Ensure output directory exists
        os.makedirs(OUTPUT_DIR, exist_ok=True)

        # Generate unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"final_invoice_{timestamp}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, output_filename)

        # Save and close the workbook
        workbook.save(output_path)
        workbook.close()

        # Ensure file exists before returning the path
        if not os.path.exists(output_path):
            raise Exception("Failed to generate Excel file")

        return output_path  # Return the file path

    except Exception as e:
        raise RuntimeError(f"Error generating Excel: {str(e)}")



# Function to convert an Excel file to PDF using ConvertAPI
def convert_excel_to_pdf(excel_file_path: str, output_pdf_path: str):
    """Converts an Excel file to a PDF using an external API"""
    try:
        headers = {"Authorization": f"Bearer {CONVERT_API_KEY}"}

        # Open and read the Excel file
        with open(excel_file_path, "rb") as file:
            files = {"File": (os.path.basename(excel_file_path), file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            data = {
                "StoreFile": "false",
                "WorksheetActive": "true",
                "PageOrientation": "landscape",
            }
            response = requests.post(CONVERT_API_URL, headers=headers, files=files, data=data, timeout=120, stream=True)

            response.raise_for_status()
            response_data = response.json()

            if "Files" not in response_data or not response_data["Files"]:
                raise Exception("No files returned in the response")

            file_data_base64 = response_data["Files"][0]["FileData"]
            file_data_bytes = base64.b64decode(file_data_base64)

            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_pdf_path), exist_ok=True)

            with open(output_pdf_path, "wb") as pdf_file:
                pdf_file.write(file_data_bytes)

            print(f"PDF successfully saved as '{output_pdf_path}'")
            return output_pdf_path

    except requests.Timeout:
        raise RuntimeError("The request to ConvertAPI timed out. Try reducing the file size.")

    except requests.RequestException as e:
        raise RuntimeError(f"Error during API request: {e}")

    except Exception as e:
        raise RuntimeError(f"Error in PDF conversion: {str(e)}")

        

def extract_pages_from_pdf(input_pdf: str, output_pdf: str, start_page: int):
    """
    Extracts pages from `start_page` to the end from `input_pdf` and saves it as `output_pdf`.
    """
    doc = fitz.open(input_pdf)
    total_pages = len(doc)

    if start_page > total_pages:
        print(f"Error: The PDF only has {total_pages} pages, cannot extract from page {start_page}")
        return False

    new_doc = fitz.open()
    for page_num in range(start_page - 1, total_pages):  # Convert to 0-based index
        new_doc.insert_pdf(doc, from_page=page_num, to_page=page_num)

    # Save to a new file instead of overwriting
    temp_output_pdf = output_pdf.replace(".pdf", "_temp.pdf")
    new_doc.save(temp_output_pdf)
    new_doc.close()
    doc.close()

    # Replace old file with new one
    os.replace(temp_output_pdf, output_pdf)

    if os.path.exists(output_pdf):
        print(f"Extracted report pages saved as '{output_pdf}'")
        return True
    else:
        print(f"Failed to save extracted report pages as '{output_pdf}'")
        return False


# Function to remove formulas from an Excel file
def remove_formulas_from_excel(input_file: str, output_file: str):
    wb = load_workbook(input_file, data_only=True)  # Load with computed values
    new_wb = load_workbook(input_file)  # Load without data_only to retain structure

    for sheet_name in wb.sheetnames:
        source_sheet = wb[sheet_name]  # Sheet with computed values
        target_sheet = new_wb[sheet_name]  # Sheet with formulas

        for row_idx, row in enumerate(source_sheet.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                target_sheet.cell(row=row_idx, column=col_idx, value=cell.value)  # Copy value, removing formula

    new_wb.save(output_file)
    print(f"Processed file saved as: {output_file}")


@app.route('/health', methods=['GET'])
def health_check():
   return jsonify({"status": "ok", "message": "Flask app is running"}), 200

# Run the Flask app
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)



